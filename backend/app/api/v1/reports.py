import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import require_accountant
from app.core.database import get_db
from app.models.order import Order, OrderItem, OrderStatus
from app.models.user import User

router = APIRouter(prefix="/reports", tags=["Reportes ERP"])


@router.get("/sales.xlsx")
async def sales_report_xlsx(
    date_from: datetime = Query(...),
    date_to: datetime = Query(...),
    user: User = Depends(require_accountant),
    db: AsyncSession = Depends(get_db),
):
    """Genera un libro Excel real (openpyxl) con el detalle de ventas por línea de pedido pagada."""
    stmt = (
        select(OrderItem, Order)
        .join(Order, Order.id == OrderItem.order_id)
        .where(
            Order.status.in_([OrderStatus.PAID, OrderStatus.PROCESSING, OrderStatus.SHIPPED, OrderStatus.DELIVERED]),
            Order.created_at >= date_from,
            Order.created_at <= date_to,
        )
        .order_by(Order.created_at)
    )
    result = await db.execute(stmt)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = ["N° Pedido", "Fecha", "SKU", "Producto", "Cantidad", "Precio Unitario", "Subtotal Línea", "Moneda", "Estado"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    total_revenue = 0.0
    for item, order in rows:
        line_total = float(item.unit_price_snapshot) * item.quantity
        total_revenue += line_total
        ws.append([
            order.order_number,
            order.created_at.strftime("%Y-%m-%d %H:%M"),
            item.sku_snapshot,
            item.product_title_snapshot,
            item.quantity,
            float(item.unit_price_snapshot),
            round(line_total, 2),
            order.currency,
            order.status.value,
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "TOTAL", round(total_revenue, 2)])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)

    for column_cells in ws.columns:
        max_length = max(len(str(c.value)) if c.value else 0 for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_ventas_{date_from.date()}_{date_to.date()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/inventory-valuation.pdf")
async def inventory_valuation_pdf(
    user: User = Depends(require_accountant),
    db: AsyncSession = Depends(get_db),
):
    """
    Genera un PDF contable real (reportlab) con la valoración de inventario
    consolidada: costo total = sum(cost_price * quantity_on_hand) por SKU.
    """
    from app.models.catalog import ProductVariant
    from app.models.inventory import InventoryItem

    stmt = (
        select(
            ProductVariant.sku,
            ProductVariant.cost_price,
            func.sum(InventoryItem.quantity_on_hand).label("total_qty"),
        )
        .join(InventoryItem, InventoryItem.variant_id == ProductVariant.id)
        .group_by(ProductVariant.id)
    )
    result = await db.execute(stmt)
    rows = result.all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Reporte de Valoración de Inventario", styles["Title"]),
        Paragraph(f"Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    table_data = [["SKU", "Costo Unitario", "Cantidad en Stock", "Valor Total"]]
    grand_total = 0.0
    for sku, cost_price, qty in rows:
        cost = float(cost_price) if cost_price else 0.0
        qty = qty or 0
        line_value = round(cost * qty, 2)
        grand_total += line_value
        table_data.append([sku, f"${cost:,.2f}", str(qty), f"${line_value:,.2f}"])

    table_data.append(["", "", "VALOR TOTAL", f"${grand_total:,.2f}"])

    table = Table(table_data, colWidths=[5 * cm, 4 * cm, 4 * cm, 4 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="valoracion_inventario.pdf"'},
    )
